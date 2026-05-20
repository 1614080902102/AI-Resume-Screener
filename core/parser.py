"""Multi-format candidate material parser.

Supports:
- PDF
- DOCX (Word)
- XLSX (Excel)
- Markdown / TXT
- Audio/Video (via Whisper)
- ZIP (containing any of the above — used for batch candidate mode)
- URL (web page)
"""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from .llm import transcribe_audio

PDF_EXTS = {".pdf"}
DOCX_EXTS = {".docx"}
XLSX_EXTS = {".xlsx", ".xls"}
TEXT_EXTS = {".md", ".markdown", ".txt"}
VIDEO_EXTS = {".mp4", ".mov", ".avi", ".webm", ".m4a", ".mp3", ".wav"}
ZIP_EXTS = {".zip"}


def parse_file(file_bytes: bytes, filename: str) -> str:
    """Parse uploaded file bytes into plain text.

    For ZIP files, all contained files are parsed and concatenated.
    """
    ext = Path(filename).suffix.lower()

    if ext in PDF_EXTS:
        return _parse_pdf(file_bytes)
    if ext in DOCX_EXTS:
        return _parse_docx(file_bytes)
    if ext in XLSX_EXTS:
        return _parse_xlsx(file_bytes)
    if ext in TEXT_EXTS:
        return file_bytes.decode("utf-8", errors="ignore")
    if ext in VIDEO_EXTS:
        return _parse_av(file_bytes, ext)
    if ext in ZIP_EXTS:
        return _parse_zip(file_bytes, filename)

    return file_bytes.decode("utf-8", errors="ignore")


def parse_url(url: str) -> str:
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        )
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        return f"[抓取失败:{url} | {e}]"
    soup = BeautifulSoup(resp.text, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = soup.get_text(separator="\n", strip=True)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return f"[来源 URL: {url}]\n{text[:8000]}"


def _parse_pdf(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    parts = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(parts).strip()


def _parse_docx(data: bytes) -> str:
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"[Sheet: {ws.title}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                out.append(" | ".join(cells))
    return "\n".join(out)


def _parse_av(data: bytes, ext: str) -> str:
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
        f.write(data)
        path = f.name
    try:
        transcript = transcribe_audio(path)
        return f"[视频/音频转写]\n{transcript}"
    finally:
        Path(path).unlink(missing_ok=True)


def _parse_zip(data: bytes, filename: str) -> str:
    """Extract all files in zip and parse them, concatenated."""
    parts = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                inner_name = info.filename
                # 忽略 macOS 元信息
                if inner_name.startswith("__MACOSX/") or inner_name.endswith(".DS_Store"):
                    continue
                try:
                    inner_bytes = zf.read(info)
                    text = parse_file(inner_bytes, inner_name)
                    parts.append(f"=== [{filename}] -> {inner_name} ===\n{text}")
                except Exception as e:
                    parts.append(f"=== [{filename}] -> {inner_name} 解析失败:{e} ===")
    except zipfile.BadZipFile as e:
        return f"[ZIP 解析失败:{e}]"
    return "\n\n".join(parts)
