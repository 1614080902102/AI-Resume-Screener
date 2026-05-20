"""HR-friendly Excel report generation."""
from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def screenings_to_excel(items: Iterable[dict], job_name: str = "") -> bytes:
    """Convert a list of screenings (from db.list_screenings) into a HR Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "候选人初筛报告"

    headers = [
        "排名", "候选人", "性别", "年龄", "现居", "工作年限",
        "匹配分", "通过", "评级", "一句话总评",
        "Top 亮点", "Top 风险", "建议追问",
        "经验分", "技能分", "领域分", "潜力分",
        "材料来源", "评估时间",
    ]

    title = f"候选人初筛报告 — {job_name}" if job_name else "候选人初筛报告"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    tcell = ws.cell(row=1, column=1, value=title)
    tcell.font = Font(size=16, bold=True, color="FFFFFF")
    tcell.fill = PatternFill("solid", fgColor="4F5BD5")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="667EEA")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    items_list = list(items)
    items_list.sort(key=lambda r: r.get("score", 0), reverse=True)

    thin = Side(border_style="thin", color="D0D7E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(items_list, 1):
        row = i + 2
        basic = (r.get("resume") or {}).get("basic", {}) or {}
        match = r.get("match") or {}
        bd = match.get("score_breakdown", {}) or {}

        def _first_str(items, key):
            if not items:
                return "—"
            it = items[0]
            if isinstance(it, dict):
                return it.get(key, "") or "—"
            return str(it)

        top_hit = _first_str(match.get("hits"), "point")
        top_risk = _first_str(match.get("risks"), "point")
        top_q = _first_str(match.get("questions"), "question")

        passed = bool(r.get("passed"))
        passed_str = "✅ 通过" if passed else "⏸ 未通过"

        values = [
            i,
            r.get("candidate_name") or basic.get("name") or "—",
            basic.get("gender") or "—",
            basic.get("age") or "—",
            basic.get("location") or "—",
            basic.get("years_of_experience") or "—",
            r.get("score") or 0,
            passed_str,
            match.get("verdict") or "—",
            match.get("summary") or "—",
            top_hit,
            top_risk,
            top_q,
            (bd.get("experience_match", {}).get("score", 0)
             if isinstance(bd.get("experience_match"), dict)
             else bd.get("experience_match") or 0),
            (bd.get("skill_match", {}).get("score", 0)
             if isinstance(bd.get("skill_match"), dict)
             else bd.get("skill_match") or 0),
            (bd.get("domain_match", {}).get("score", 0)
             if isinstance(bd.get("domain_match"), dict)
             else bd.get("domain_match") or 0),
            (bd.get("potential", {}).get("score", 0)
             if isinstance(bd.get("potential"), dict)
             else bd.get("potential") or 0),
            r.get("candidate_source") or "—",
            (r.get("created_at") or "")[:16],
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == 7:  # 匹配分
                cell.font = Font(bold=True, size=12)
            if c == 8:  # 通过
                if passed:
                    cell.fill = PatternFill("solid", fgColor="D4EDDA")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFF3CD")

        # 按通过/不通过给整行加底色
        row_fill = "F0FAF3" if passed else "FFFFFF"
        for c in range(1, len(values) + 1):
            cur = ws.cell(row=row, column=c)
            if c != 8 and not cur.fill.start_color.rgb:
                cur.fill = PatternFill("solid", fgColor=row_fill)

    # 列宽
    col_widths = {
        1: 5, 2: 12, 3: 6, 4: 6, 5: 10, 6: 8, 7: 9, 8: 10, 9: 16, 10: 28,
        11: 38, 12: 38, 13: 38,
        14: 7, 15: 7, 16: 7, 17: 7,
        18: 28, 19: 18,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
